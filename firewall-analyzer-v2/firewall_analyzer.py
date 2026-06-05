#!/usr/bin/env python3
"""
firewall_analyzer.py - Schema-flexible firewall log analyzer, threat detector,
and firewall-rule generator.

The tool ingests a CSV of firewall/connection logs in *any* reasonable column
layout, normalizes it to a common schema, summarizes the traffic, flags likely
malicious or suspicious patterns, and proposes concrete ALLOW / BLOCK rules in
generic, iptables, and pfSense (pf) syntax.

Design goals
------------
* Zero required third-party dependencies. The core runs on the Python standard
  library alone, so `python3 firewall_analyzer.py logs.csv` works anywhere.
* Optional enhancements degrade gracefully:
    - `openpyxl`  -> a multi-sheet .xlsx workbook (`--excel`)
    - `matplotlib`-> PNG charts (`--charts`)
* Schema flexibility. Column names are matched against an alias table and, when
  that fails, inferred from the data itself (values that look like IPs, ports,
  actions, or timestamps).

Author: Scott Smith  (CSC-842)
License: MIT
"""

from __future__ import annotations

import argparse
import csv
import ipaddress
import json
import logging
import os
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

__version__ = "2.0.0"

log = logging.getLogger("firewall_analyzer")


# ---------------------------------------------------------------------------
# Reference data
# ---------------------------------------------------------------------------

# port -> (service name, default transport protocol)
WELL_KNOWN_PORTS: Dict[int, Tuple[str, str]] = {
    20: ("ftp-data", "tcp"), 21: ("ftp", "tcp"), 22: ("ssh", "tcp"),
    23: ("telnet", "tcp"), 25: ("smtp", "tcp"), 53: ("dns", "udp"),
    67: ("dhcp", "udp"), 68: ("dhcp", "udp"), 69: ("tftp", "udp"),
    80: ("http", "tcp"), 88: ("kerberos", "tcp"), 110: ("pop3", "tcp"),
    111: ("rpcbind", "tcp"), 123: ("ntp", "udp"), 135: ("msrpc", "tcp"),
    137: ("netbios-ns", "udp"), 138: ("netbios-dgm", "udp"),
    139: ("netbios-ssn", "tcp"), 143: ("imap", "tcp"), 161: ("snmp", "udp"),
    162: ("snmptrap", "udp"), 179: ("bgp", "tcp"), 389: ("ldap", "tcp"),
    443: ("https", "tcp"), 445: ("smb", "tcp"), 465: ("smtps", "tcp"),
    500: ("isakmp", "udp"), 514: ("syslog", "udp"), 515: ("printer", "tcp"),
    587: ("submission", "tcp"), 631: ("ipp", "tcp"), 636: ("ldaps", "tcp"),
    873: ("rsync", "tcp"), 993: ("imaps", "tcp"), 995: ("pop3s", "tcp"),
    1080: ("socks", "tcp"), 1194: ("openvpn", "udp"), 1433: ("mssql", "tcp"),
    1434: ("mssql-m", "udp"), 1521: ("oracle", "tcp"), 1701: ("l2tp", "udp"),
    1723: ("pptp", "tcp"), 1812: ("radius", "udp"), 2049: ("nfs", "tcp"),
    2082: ("cpanel", "tcp"), 2222: ("ssh-alt", "tcp"), 3128: ("squid", "tcp"),
    3268: ("ldap-gc", "tcp"), 3306: ("mysql", "tcp"), 3389: ("rdp", "tcp"),
    4444: ("metasploit", "tcp"), 5060: ("sip", "udp"), 5432: ("postgresql", "tcp"),
    5555: ("freeciv/adb", "tcp"), 5601: ("kibana", "tcp"), 5900: ("vnc", "tcp"),
    5985: ("winrm-http", "tcp"), 5986: ("winrm-https", "tcp"),
    6379: ("redis", "tcp"), 6443: ("kubernetes-api", "tcp"),
    8000: ("http-alt", "tcp"), 8080: ("http-proxy", "tcp"),
    8443: ("https-alt", "tcp"), 8888: ("http-alt", "tcp"),
    9000: ("http-alt", "tcp"), 9200: ("elasticsearch", "tcp"),
    9300: ("elasticsearch-transport", "tcp"), 11211: ("memcached", "tcp"),
    27017: ("mongodb", "tcp"), 27018: ("mongodb", "tcp"),
    50070: ("hadoop", "tcp"),
}

# Ports that are high value to an attacker and rarely belong open to the world.
# Inbound external -> internal hits on these are treated as high risk.
SENSITIVE_PORTS: Dict[int, str] = {
    21: "FTP (cleartext)", 23: "Telnet (cleartext)", 135: "MS RPC",
    137: "NetBIOS", 138: "NetBIOS", 139: "NetBIOS/SMB", 445: "SMB",
    1433: "MS-SQL", 1521: "Oracle DB", 2049: "NFS", 3306: "MySQL",
    3389: "RDP", 5432: "PostgreSQL", 5900: "VNC", 5985: "WinRM",
    5986: "WinRM", 6379: "Redis (no auth by default)",
    9200: "Elasticsearch", 11211: "Memcached", 27017: "MongoDB",
}

# Ports an internal host should rarely dial OUT to. Outbound hits here are
# treated as likely command-and-control / tooling regardless of the well-known
# table (4444 is metasploit's default, 1080/9050 are proxies/Tor, etc.).
KNOWN_BAD_EGRESS_PORTS: Dict[int, str] = {
    1080: "SOCKS proxy", 4444: "Metasploit default", 4445: "Metasploit alt",
    5555: "ADB/remote", 6660: "IRC", 6667: "IRC", 6697: "IRC over TLS",
    9001: "Tor ORPort", 9050: "Tor SOCKS", 9051: "Tor control",
    12345: "NetBus", 31337: "Back Orifice",
}

# Destination ports that are normal for outbound client traffic; egress to
# these is not, by itself, suspicious.
COMMON_BENIGN_EGRESS_PORTS = {
    20, 21, 22, 25, 53, 80, 110, 123, 143, 443, 465, 587, 853, 993, 995,
    8080, 8443,
}

# Raw action strings, normalized -> ALLOW / BLOCK.
ALLOW_TOKENS = {"allow", "allowed", "permit", "permitted", "accept",
                "accepted", "pass", "passed", "0", "ok"}
BLOCK_TOKENS = {"deny", "denied", "drop", "dropped", "block", "blocked",
                "reject", "rejected", "1", "fail", "failed"}

# Canonical field -> set of header aliases (compared after normalization).
HEADER_ALIASES: Dict[str, set] = {
    "src_ip": {"srcip", "sourceip", "source", "src", "sourceaddress",
               "srcaddr", "saddr", "sourceaddr", "clientip", "origin",
               "srcipaddr", "sourceipaddress", "from", "sourcehost"},
    "dst_ip": {"dstip", "destip", "destinationip", "destination", "dst",
               "destaddr", "daddr", "destinationaddr", "targetip",
               "serverip", "dest", "destinationipaddress", "to", "desthost"},
    "src_port": {"srcport", "sourceport", "sport", "spt", "sourceports",
                 "srcprt", "clientport"},
    "dst_port": {"dstport", "destport", "destinationport", "dport", "dpt",
                 "service", "targetport", "destprt", "serverport", "port"},
    "protocol": {"protocol", "proto", "ipproto", "transport", "ipprotocol",
                 "l4proto"},
    "action": {"action", "disposition", "verdict", "decision", "status",
               "filteraction", "act", "fwaction", "result", "ruleaction",
               "policyaction"},
    "timestamp": {"timestamp", "time", "datetime", "date", "eventtime",
                  "starttime", "ts", "timestamp", "receivedtime", "logtime",
                  "firstpacket", "recordtime", "atimestamp"},
    "bytes": {"bytes", "bytecount", "totalbytes", "length", "len", "size",
              "datalen", "bytessent", "bytesreceived"},
    "packets": {"packets", "pkts", "packetcount", "pktcount", "npackets"},
    "interface": {"interface", "iface", "inint", "ifname", "ininterface",
                  "ingressinterface"},
    "rule": {"rule", "ruleid", "policyid", "rulename", "policy", "policyname",
             "signature", "sid"},
}

TIMESTAMP_FORMATS = (
    "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M",
    "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%d/%m/%Y %H:%M:%S",
    "%b %d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d",
)


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def _normalize_header(name: str) -> str:
    """Lowercase and strip non-alphanumerics so 'Source IP' == 'source_ip'."""
    return "".join(ch for ch in name.lower() if ch.isalnum())


def _parse_ip(value: str) -> Optional[ipaddress._BaseAddress]:
    try:
        return ipaddress.ip_address(value.strip())
    except (ValueError, AttributeError):
        return None


def _parse_port(value: str) -> Optional[int]:
    try:
        p = int(str(value).strip())
    except (ValueError, TypeError):
        return None
    return p if 0 <= p <= 65535 else None


def _parse_timestamp(value: str) -> Optional[datetime]:
    if not value:
        return None
    value = value.strip()
    # Native ISO 8601 first (handles fractional seconds / offsets).
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        pass
    # Unix epoch (seconds or milliseconds).
    if value.isdigit():
        n = int(value)
        try:
            return datetime.fromtimestamp(n / 1000 if n > 10_000_000_000 else n)
        except (ValueError, OSError, OverflowError):
            return None
    for fmt in TIMESTAMP_FORMATS:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _is_internal(ip: Optional[ipaddress._BaseAddress]) -> bool:
    """Treat RFC1918 / loopback / link-local / unique-local as 'internal'."""
    if ip is None:
        return False
    return bool(ip.is_private or ip.is_loopback or ip.is_link_local)


def service_for_port(port: Optional[int], proto: Optional[str] = None) -> str:
    if port is None:
        return "unknown"
    name, _ = WELL_KNOWN_PORTS.get(port, ("", ""))
    return name or "ephemeral/other"


def proto_for_port(port: Optional[int], explicit: Optional[str]) -> str:
    if explicit:
        return explicit.lower()
    if port is not None and port in WELL_KNOWN_PORTS:
        return WELL_KNOWN_PORTS[port][1]
    return "tcp"  # safe default for rule rendering


# ---------------------------------------------------------------------------
# Schema detection
# ---------------------------------------------------------------------------

def detect_schema(headers: Sequence[str],
                  sample_rows: Sequence[Sequence[str]]) -> Dict[str, str]:
    """
    Map canonical field names -> actual header strings.

    Step 1: alias matching on normalized header names.
    Step 2: value-based heuristics for any canonical field still unmapped,
            so the tool tolerates unfamiliar / vendor-specific column names.
    """
    mapping: Dict[str, str] = {}
    norm = {h: _normalize_header(h) for h in headers}

    # Step 1 - alias match.
    for canonical, aliases in HEADER_ALIASES.items():
        for header in headers:
            if norm[header] in aliases:
                mapping[canonical] = header
                break

    # Build per-column value profiles for heuristic inference.
    cols: Dict[str, List[str]] = {h: [] for h in headers}
    for row in sample_rows:
        for h, val in zip(headers, row):
            if val not in ("", None):
                cols[h].append(val)

    used = set(mapping.values())

    def profile(values: List[str]) -> Dict[str, float]:
        if not values:
            return {"ip": 0, "port": 0, "action": 0, "time": 0}
        n = len(values)
        ip = sum(1 for v in values if _parse_ip(v)) / n
        port = sum(1 for v in values if _parse_port(v) is not None) / n
        action = sum(1 for v in values
                     if v.strip().lower() in ALLOW_TOKENS | BLOCK_TOKENS) / n
        time = sum(1 for v in values if _parse_timestamp(v)) / n
        return {"ip": ip, "port": port, "action": action, "time": time}

    profiles = {h: profile(vals) for h, vals in cols.items()}

    # Step 2 - heuristic fill for the important fields.
    def candidates(kind: str, threshold: float) -> List[str]:
        scored = [(h, profiles[h][kind]) for h in headers
                  if h not in used and profiles[h][kind] >= threshold]
        # Preserve original column order among ties (left-to-right).
        return [h for h, _ in scored]

    # IP columns: first unused IP-like column -> src, second -> dst.
    if "src_ip" not in mapping or "dst_ip" not in mapping:
        ip_cols = candidates("ip", 0.7)
        if "src_ip" not in mapping and ip_cols:
            mapping["src_ip"] = ip_cols.pop(0)
            used.add(mapping["src_ip"])
        if "dst_ip" not in mapping and ip_cols:
            mapping["dst_ip"] = ip_cols.pop(0)
            used.add(mapping["dst_ip"])

    # Port columns: ports are integers 0-65535 but NOT IPs.
    if "src_port" not in mapping or "dst_port" not in mapping:
        port_cols = [h for h in candidates("port", 0.8)
                     if profiles[h]["ip"] < 0.5]
        if "dst_port" not in mapping and port_cols:
            # destination port is usually the more "meaningful" one; if only
            # one port column exists, treat it as destination.
            mapping["dst_port"] = port_cols.pop(-1)
            used.add(mapping["dst_port"])
        if "src_port" not in mapping and port_cols:
            mapping["src_port"] = port_cols.pop(0)
            used.add(mapping["src_port"])

    if "action" not in mapping:
        act = candidates("action", 0.6)
        if act:
            mapping["action"] = act[0]
            used.add(act[0])

    if "timestamp" not in mapping:
        tm = [h for h in candidates("time", 0.7) if profiles[h]["port"] < 0.5]
        if tm:
            mapping["timestamp"] = tm[0]
            used.add(tm[0])

    return mapping


def normalize_action(raw: Optional[str]) -> str:
    if raw is None:
        return "UNKNOWN"
    token = raw.strip().lower()
    if token in ALLOW_TOKENS:
        return "ALLOW"
    if token in BLOCK_TOKENS:
        return "BLOCK"
    # substring fallback (e.g. "Drop (out of state)")
    if any(t in token for t in ("deny", "drop", "block", "reject", "fail")):
        return "BLOCK"
    if any(t in token for t in ("allow", "permit", "accept", "pass")):
        return "ALLOW"
    return "UNKNOWN"


# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

@dataclass
class Record:
    src_ip: Optional[str] = None
    src_port: Optional[int] = None
    dst_ip: Optional[str] = None
    dst_port: Optional[int] = None
    protocol: Optional[str] = None
    action: str = "UNKNOWN"
    raw_action: Optional[str] = None
    timestamp: Optional[datetime] = None
    nbytes: Optional[int] = None

    @property
    def src_internal(self) -> bool:
        return _is_internal(_parse_ip(self.src_ip or ""))

    @property
    def dst_internal(self) -> bool:
        return _is_internal(_parse_ip(self.dst_ip or ""))


@dataclass
class Finding:
    title: str
    severity: str                 # CRITICAL / HIGH / MEDIUM / LOW / INFO
    confidence: str               # high / medium / low
    description: str              # plain-English interpretation
    recommendation: str
    entities: Dict[str, Any] = field(default_factory=dict)
    suggested_rules: List["Rule"] = field(default_factory=list)


@dataclass
class Rule:
    decision: str                 # ALLOW / BLOCK
    src: str
    dst: str
    port: Optional[int]
    proto: str
    reason: str = ""
    count: int = 0                # connections behind the rule (for sorting)


@dataclass
class AnalysisConfig:
    """All tunables in one place so the CLI and GUI share a single code path."""
    rule_format: str = "generic"          # generic | iptables | pf
    subnet_mode: str = "adaptive"         # adaptive | always | never
    subnet_min_hosts: int = 3             # adaptive: hosts/subnet before collapse
    prefix_v4: int = 24
    prefix_v6: int = 64
    top_rules: int = 25
    scan_port_threshold: int = 15
    sweep_host_threshold: int = 10
    block_threshold: int = 5
    block_ratio: float = 0.5


@dataclass
class AnalysisResult:
    records: List["Record"]
    mapping: Dict[str, str]
    summary: Dict[str, Any]
    findings: List["Finding"]
    allow_rules: List["Rule"]
    block_rules: List["Rule"]
    config: "AnalysisConfig"


SEVERITY_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}


# ---------------------------------------------------------------------------
# Ingestion
# ---------------------------------------------------------------------------

def load_records(path: str) -> Tuple[List[Record], Dict[str, str], List[str]]:
    """Read the CSV, detect its schema, and return normalized Records."""
    with open(path, "r", newline="", encoding="utf-8-sig", errors="replace") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(fh, dialect)
        try:
            headers = next(reader)
        except StopIteration:
            raise ValueError("CSV file is empty.")
        raw_rows = [row for row in reader if any(c.strip() for c in row)]

    if not raw_rows:
        raise ValueError("CSV contains a header but no data rows.")

    mapping = detect_schema(headers, raw_rows[:200])
    if "src_ip" not in mapping or "dst_ip" not in mapping:
        raise ValueError(
            "Could not identify source and destination IP columns.\n"
            f"  Detected headers: {headers}\n"
            f"  Detected mapping: {mapping}\n"
            "  Rename columns or check the file format."
        )

    idx = {canon: headers.index(col) for canon, col in mapping.items()}

    def cell(row: Sequence[str], canon: str) -> Optional[str]:
        i = idx.get(canon)
        if i is None or i >= len(row):
            return None
        v = row[i].strip()
        return v or None

    records: List[Record] = []
    for row in raw_rows:
        raw_action = cell(row, "action")
        rec = Record(
            src_ip=cell(row, "src_ip"),
            dst_ip=cell(row, "dst_ip"),
            src_port=_parse_port(cell(row, "src_port") or ""),
            dst_port=_parse_port(cell(row, "dst_port") or ""),
            protocol=(cell(row, "protocol") or None),
            raw_action=raw_action,
            action=normalize_action(raw_action),
            timestamp=_parse_timestamp(cell(row, "timestamp") or ""),
            nbytes=_to_int(cell(row, "bytes")),
        )
        if rec.src_ip and rec.dst_ip:
            records.append(rec)

    return records, mapping, headers


def _to_int(value: Optional[str]) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Traffic summary
# ---------------------------------------------------------------------------

def summarize(records: List[Record]) -> Dict[str, Any]:
    has_action = any(r.action != "UNKNOWN" for r in records)
    has_time = any(r.timestamp for r in records)

    allowed = sum(1 for r in records if r.action == "ALLOW")
    blocked = sum(1 for r in records if r.action == "BLOCK")

    top_dst_ports = Counter(r.dst_port for r in records if r.dst_port is not None)
    top_talkers = Counter(r.src_ip for r in records)
    top_targets = Counter(r.dst_ip for r in records)
    protocols = Counter((r.protocol or proto_for_port(r.dst_port, None))
                        for r in records)

    timespan = None
    if has_time:
        times = sorted(r.timestamp for r in records if r.timestamp)
        if times:
            timespan = (times[0], times[-1])

    external_inbound = sum(1 for r in records
                           if not r.src_internal and r.dst_internal)
    internal_outbound = sum(1 for r in records
                            if r.src_internal and not r.dst_internal)

    return {
        "total": len(records),
        "has_action": has_action,
        "has_time": has_time,
        "allowed": allowed,
        "blocked": blocked,
        "unknown_action": len(records) - allowed - blocked,
        "unique_sources": len(top_talkers),
        "unique_targets": len(top_targets),
        "external_inbound": external_inbound,
        "internal_outbound": internal_outbound,
        "top_dst_ports": top_dst_ports.most_common(15),
        "top_talkers": top_talkers.most_common(15),
        "top_targets": top_targets.most_common(15),
        "protocols": protocols.most_common(),
        "timespan": timespan,
    }


# ---------------------------------------------------------------------------
# Threat detection
# ---------------------------------------------------------------------------

def detect_threats(records: List[Record],
                   scan_port_threshold: int = 15,
                   sweep_host_threshold: int = 10,
                   block_threshold: int = 5,
                   block_ratio: float = 0.5) -> List[Finding]:
    findings: List[Finding] = []

    scans = _detect_vertical_scan(records, scan_port_threshold)
    sweeps = _detect_horizontal_sweep(records, sweep_host_threshold)
    sensitive = _detect_sensitive_inbound(records)
    blocks = _detect_high_block_sources(records, block_threshold, block_ratio)
    egress = _detect_suspicious_egress(records)

    # Correlation: if an external source is already flagged for a scan or a
    # sweep, fold its individual sensitive-inbound hits into that broader
    # finding instead of emitting dozens of near-duplicate CRITICALs.
    scan_sources = {f.entities.get("source") for f in scans}
    sweep_src_ports = {(f.entities.get("source"), f.entities.get("port"))
                       for f in sweeps}
    deduped_sensitive = []
    folded = 0
    for f in sensitive:
        src = f.entities.get("source")
        port = f.entities.get("port")
        if src in scan_sources or (src, port) in sweep_src_ports:
            folded += 1
            continue
        deduped_sensitive.append(f)
    if folded:
        log.debug("Folded %d sensitive-inbound hits into scan/sweep findings",
                  folded)

    findings = scans + sweeps + deduped_sensitive + blocks + egress
    findings.sort(key=lambda f: (SEVERITY_ORDER.get(f.severity, 9), f.title))
    return findings


def _detect_vertical_scan(records, threshold: int) -> List[Finding]:
    """One source touching many distinct ports on one destination = port scan."""
    by_pair: Dict[Tuple[str, str], set] = defaultdict(set)
    for r in records:
        if r.dst_port is not None:
            by_pair[(r.src_ip, r.dst_ip)].add(r.dst_port)
    out = []
    for (src, dst), ports in by_pair.items():
        if len(ports) >= threshold:
            external = not _is_internal(_parse_ip(src))
            sev = "HIGH" if external else "MEDIUM"
            sample = sorted(ports)[:20]
            out.append(Finding(
                title=f"Port scan: {src} -> {dst}",
                severity=sev,
                confidence="high",
                description=(
                    f"{src} connected to {len(ports)} distinct destination "
                    f"ports on {dst}. Hitting many ports on a single host is "
                    "the classic signature of a vertical port scan / service "
                    "enumeration."
                    + ("" if external else
                       " The source is internal, so this may also be a "
                       "vulnerability scanner or a compromised host.")
                ),
                recommendation=(
                    f"Investigate {src}. If it is not an authorized scanner, "
                    f"block it at the perimeter."
                ),
                entities={"source": src, "target": dst,
                          "distinct_ports": len(ports), "sample_ports": sample},
                suggested_rules=[Rule("BLOCK", src, "any", None, "tcp",
                                      "vertical port scan")],
            ))
    return out


def _detect_horizontal_sweep(records, threshold: int) -> List[Finding]:
    """One source hitting the same port across many hosts = network sweep."""
    by_srcport: Dict[Tuple[str, int], set] = defaultdict(set)
    for r in records:
        if r.dst_port is not None:
            by_srcport[(r.src_ip, r.dst_port)].add(r.dst_ip)
    out = []
    for (src, port), hosts in by_srcport.items():
        if len(hosts) >= threshold:
            external = not _is_internal(_parse_ip(src))
            svc = service_for_port(port)
            sensitive = port in SENSITIVE_PORTS
            sev = "HIGH" if (external or sensitive) else "MEDIUM"
            out.append(Finding(
                title=f"Network sweep: {src} -> *:{port}",
                severity=sev,
                confidence="high",
                description=(
                    f"{src} contacted {len(hosts)} distinct hosts all on port "
                    f"{port} ({svc}). Fanning out across many hosts on one "
                    "port is a horizontal sweep, typically reconnaissance for "
                    "a specific exploitable service"
                    + (f" ({SENSITIVE_PORTS[port]})." if sensitive else ".")
                ),
                recommendation=(
                    f"Confirm whether {src} is an approved scanner. If not, "
                    "treat as reconnaissance and block."
                ),
                entities={"source": src, "port": port, "service": svc,
                          "distinct_hosts": len(hosts)},
                suggested_rules=[Rule("BLOCK", src, "any", None, "tcp",
                                      f"horizontal sweep on port {port}")],
            ))
    return out


def _detect_sensitive_inbound(records) -> List[Finding]:
    """External hosts reaching internal sensitive services."""
    hits: Dict[Tuple[str, str, int], int] = Counter()
    for r in records:
        if (not r.src_internal and r.dst_internal
                and r.dst_port in SENSITIVE_PORTS
                and r.action != "BLOCK"):
            hits[(r.src_ip, r.dst_ip, r.dst_port)] += 1
    out = []
    for (src, dst, port), count in sorted(hits.items(),
                                          key=lambda kv: -kv[1]):
        svc = SENSITIVE_PORTS[port]
        out.append(Finding(
            title=f"External access to sensitive service: {dst}:{port}",
            severity="CRITICAL",
            confidence="high",
            description=(
                f"External host {src} reached internal {dst} on port {port} "
                f"({svc}){' ' + str(count) + ' times' if count > 1 else ''}. "
                "Management and database services should not be exposed to "
                "untrusted networks; this is a common initial-access and "
                "lateral-movement vector."
            ),
            recommendation=(
                f"Restrict port {port} on {dst} to known internal management "
                "ranges or a VPN/jump host, and confirm the traffic was not "
                "an intrusion."
            ),
            entities={"source": src, "target": dst, "port": port,
                      "service": svc, "count": count},
            suggested_rules=[
                Rule("BLOCK", src, dst, port,
                     proto_for_port(port, None),
                     f"external access to {svc}")],
        ))
    return out


def _detect_high_block_sources(records, min_blocks: int,
                               min_ratio: float) -> List[Finding]:
    """Sources generating many blocked events (brute force / scanning)."""
    if not any(r.action == "BLOCK" for r in records):
        return []
    totals: Counter = Counter()
    blocks: Counter = Counter()
    for r in records:
        totals[r.src_ip] += 1
        if r.action == "BLOCK":
            blocks[r.src_ip] += 1
    out = []
    for src, nblock in blocks.items():
        ratio = nblock / totals[src]
        if nblock >= min_blocks and ratio >= min_ratio:
            external = not _is_internal(_parse_ip(src))
            out.append(Finding(
                title=f"High block volume from {src}",
                severity="HIGH" if external else "MEDIUM",
                confidence="medium",
                description=(
                    f"{src} generated {nblock} blocked events out of "
                    f"{totals[src]} ({ratio:.0%} blocked). A source that the "
                    "firewall repeatedly denies is usually probing for an open "
                    "path, brute forcing, or misconfigured."
                ),
                recommendation=(
                    f"Review what {src} is attempting. Sustained denied "
                    "traffic from one external source warrants an explicit "
                    "perimeter block and, if internal, host investigation."
                ),
                entities={"source": src, "blocked": nblock,
                          "total": totals[src], "block_ratio": round(ratio, 3)},
                suggested_rules=[Rule("BLOCK", src, "any", None, "tcp",
                                      f"{nblock} blocked events")],
            ))
    out.sort(key=lambda f: -f.entities["blocked"])
    return out


def _detect_suspicious_egress(records) -> List[Finding]:
    """Internal hosts beaconing out to external IPs on uncommon or known-bad
    ports. Known-bad ports (e.g. 4444, Tor, IRC) are raised to MEDIUM."""
    egress: Dict[Tuple[str, str, int], int] = Counter()
    for r in records:
        if not (r.src_internal and not r.dst_internal):
            continue
        if r.dst_port is None or r.action == "BLOCK":
            continue
        known_bad = r.dst_port in KNOWN_BAD_EGRESS_PORTS
        uncommon = (r.dst_port not in COMMON_BENIGN_EGRESS_PORTS
                    and r.dst_port > 1024)
        if known_bad or uncommon:
            egress[(r.src_ip, r.dst_ip, r.dst_port)] += 1

    out = []
    for (src, dst, port), count in sorted(egress.items(),
                                          key=lambda kv: -kv[1])[:25]:
        known_bad = port in KNOWN_BAD_EGRESS_PORTS
        sev = "MEDIUM" if known_bad else "LOW"
        conf = "medium" if known_bad else "low"
        tag = (f" Port {port} is associated with "
               f"{KNOWN_BAD_EGRESS_PORTS[port]}." if known_bad else "")
        out.append(Finding(
            title=f"Suspicious outbound: {src} -> {dst}:{port}",
            severity=sev,
            confidence=conf,
            description=(
                f"Internal host {src} sent outbound traffic to external {dst} "
                f"on port {port}"
                f"{' (' + str(count) + ' connections)' if count > 1 else ''}."
                f"{tag} Outbound connections to uncommon or known-bad ports "
                "can be benign (custom apps) but also match command-and-control "
                "or data-exfiltration behavior."
            ),
            recommendation=(
                f"Verify the application on {src} using port {port}. If "
                "unexplained, inspect the destination reputation and the host "
                "itself. Known-bad ports warrant priority review."
            ),
            entities={"source": src, "target": dst, "port": port,
                      "count": count, "known_bad": known_bad},
            suggested_rules=([Rule("BLOCK", src, dst, port,
                                   proto_for_port(port, None),
                                   f"egress to {KNOWN_BAD_EGRESS_PORTS[port]} "
                                   f"port")] if known_bad else []),
        ))
    return out


# ---------------------------------------------------------------------------
# Rule generation
# ---------------------------------------------------------------------------

def _subnet_of(ip: str, prefix_v4: int,
               prefix_v6: int) -> Optional[ipaddress._BaseNetwork]:
    """Return the containing network for an IP at the configured prefix."""
    addr = _parse_ip(ip)
    if isinstance(addr, ipaddress.IPv4Address):
        return ipaddress.ip_network(f"{ip}/{prefix_v4}", strict=False)
    if isinstance(addr, ipaddress.IPv6Address):
        return ipaddress.ip_network(f"{ip}/{prefix_v6}", strict=False)
    return None


def build_allow_rules(records: List[Record], findings: List[Finding],
                      config: AnalysisConfig) -> List[Rule]:
    """
    Propose ALLOW rules from established, non-malicious flows.

    A flow is eligible if the firewall did not block it (when action is known)
    and its source is not implicated in any threat finding.

    Subnet aggregation collapses many hosts in the same subnet that talk to the
    same destination+port into a single CIDR rule. This is the realistic shape
    of firewall policy and keeps the output manageable on large captures:

      * adaptive (default): collapse a subnet only when at least
        ``subnet_min_hosts`` distinct hosts from it hit the same dst+port;
        otherwise emit per-host rules. So a subnet with one busy host stays
        specific, while a subnet with a dozen clients becomes one rule.
      * always: collapse every flow to its source subnet.
      * never: one rule per individual source host (legacy behavior).
    """
    flagged_sources = {
        f.entities.get("source")
        for f in findings if f.entities.get("source")
    }

    # (dst, port, proto) -> {src_ip: connection_count}
    flows: Dict[Tuple[str, int, str], Counter] = defaultdict(Counter)
    for r in records:
        if r.action == "BLOCK" or r.dst_port is None:
            continue
        if r.src_ip in flagged_sources:
            continue
        proto = proto_for_port(r.dst_port, r.protocol)
        flows[(r.dst_ip, r.dst_port, proto)][r.src_ip] += 1

    rules: List[Rule] = []
    for (dst, port, proto), src_counter in flows.items():
        # Bucket the sources of this flow by their subnet.
        buckets: Dict[Any, List[str]] = defaultdict(list)
        for src in src_counter:
            net = _subnet_of(src, config.prefix_v4, config.prefix_v6)
            buckets[net if net is not None else src].append(src)

        for net, srcs in buckets.items():
            conns = sum(src_counter[s] for s in srcs)
            is_net = isinstance(net, ipaddress._BaseNetwork)
            collapse = is_net and (
                config.subnet_mode == "always"
                or (config.subnet_mode == "adaptive"
                    and len(srcs) >= config.subnet_min_hosts)
            )
            if collapse:
                rules.append(Rule(
                    "ALLOW", str(net), dst, port, proto,
                    reason=(f"{len(srcs)} hosts in subnet, {conns} connections "
                            f"({service_for_port(port)})"),
                    count=conns))
            else:
                for s in srcs:
                    rules.append(Rule(
                        "ALLOW", s, dst, port, proto,
                        reason=(f"{src_counter[s]} connections "
                                f"({service_for_port(port)})"),
                        count=src_counter[s]))

    rules.sort(key=lambda r: (-r.count, r.dst, r.port or 0))
    return rules[:config.top_rules]


def build_block_rules(findings: List[Finding],
                      config: AnalysisConfig) -> List[Rule]:
    """
    Collect BLOCK rules from findings, de-duplicate, and consolidate.

    Broad source blocks (block this IP entirely) are consolidated to a subnet
    block when two or more flagged sources share the same subnet -- e.g. a
    distributed scan from one /24 becomes a single block rule. Precise rules
    (block a specific destination/port) are kept as-is.
    """
    seen = set()
    raw: List[Rule] = []
    for f in findings:
        for rule in f.suggested_rules:
            if rule.decision != "BLOCK":
                continue
            key = (rule.src, rule.dst, rule.port)
            if key in seen:
                continue
            seen.add(key)
            raw.append(rule)

    broad = [r for r in raw if r.dst == "any" and r.port is None]
    specific = [r for r in raw if not (r.dst == "any" and r.port is None)]

    if config.subnet_mode == "never":
        return specific + broad

    by_net: Dict[Any, List[Rule]] = defaultdict(list)
    for r in broad:
        net = _subnet_of(r.src, config.prefix_v4, config.prefix_v6)
        by_net[net if net is not None else r.src].append(r)

    consolidated: List[Rule] = []
    for net, rs in by_net.items():
        distinct = {r.src for r in rs}
        is_net = isinstance(net, ipaddress._BaseNetwork)
        if is_net and (config.subnet_mode == "always" or len(distinct) >= 2):
            consolidated.append(Rule(
                "BLOCK", str(net), "any", None, "tcp",
                reason=f"{len(distinct)} flagged source(s) in subnet"))
        else:
            consolidated.extend(rs)

    return specific + consolidated


# --- rule renderers --------------------------------------------------------

def render_rule(rule: Rule, fmt: str) -> str:
    if fmt == "iptables":
        return _render_iptables(rule)
    if fmt == "pf":
        return _render_pf(rule)
    return _render_generic(rule)


def _render_generic(r: Rule) -> str:
    port = f":{r.port}/{r.proto}" if r.port is not None else " (all ports)"
    base = f"{r.decision:5s} {r.src} -> {r.dst}{port}"
    return f"{base}    # {r.reason}" if r.reason else base


def _render_iptables(r: Rule) -> str:
    target = "ACCEPT" if r.decision == "ALLOW" else "DROP"
    parts = ["iptables", "-A", "FORWARD"]
    if r.src != "any":
        parts += ["-s", r.src]
    if r.dst != "any":
        parts += ["-d", r.dst]
    if r.port is not None:
        parts += ["-p", r.proto, "--dport", str(r.port)]
    parts += ["-j", target]
    line = " ".join(parts)
    return f"{line}    # {r.reason}" if r.reason else line


def _render_pf(r: Rule) -> str:
    # pfSense / OpenBSD pf syntax.
    action = "pass" if r.decision == "ALLOW" else "block return"
    src = "any" if r.src == "any" else r.src
    dst = "any" if r.dst == "any" else r.dst
    proto = f"proto {r.proto} " if r.port is not None else ""
    port = f" port {r.port}" if r.port is not None else ""
    keep = " keep state" if r.decision == "ALLOW" else ""
    line = f"{action} in quick {proto}from {src} to {dst}{port}{keep}"
    return f"{line}    # {r.reason}" if r.reason else line


# ---------------------------------------------------------------------------
# Reporting
# ---------------------------------------------------------------------------

def format_report(result: "AnalysisResult") -> str:
    """Render the full analysis as a plain-text report string."""
    mapping = result.mapping
    summary = result.summary
    findings = result.findings
    rule_fmt = result.config.rule_format
    out: List[str] = []
    w = out.append
    bar = "=" * 70

    w(f"\n{bar}\n  FIREWALL LOG ANALYSIS REPORT\n{bar}\n")
    w(f"  Records analyzed : {summary['total']}\n")
    w(f"  Unique sources   : {summary['unique_sources']}\n")
    w(f"  Unique targets   : {summary['unique_targets']}\n")
    if summary["timespan"]:
        start, end = summary["timespan"]
        w(f"  Time span        : {start}  ->  {end}\n")
    if summary["has_action"]:
        w(f"  Allowed / Blocked: {summary['allowed']} / {summary['blocked']}")
        if summary["unknown_action"]:
            w(f"  ({summary['unknown_action']} unknown)")
        w("\n")
    w(f"  External inbound : {summary['external_inbound']}   "
      f"Internal outbound: {summary['internal_outbound']}\n")

    w("\n  Detected columns:\n")
    for canon, col in sorted(mapping.items()):
        w(f"    {canon:11s} <- '{col}'\n")

    w(f"\n{bar}\n  TOP DESTINATION PORTS\n{bar}\n")
    for port, count in summary["top_dst_ports"]:
        w(f"  {count:6d}  port {port:<6d} {service_for_port(port)}\n")

    w(f"\n{bar}\n  TOP SOURCES (talkers)\n{bar}\n")
    for ip, count in summary["top_talkers"]:
        tag = "internal" if _is_internal(_parse_ip(ip)) else "EXTERNAL"
        w(f"  {count:6d}  {ip:<18s} [{tag}]\n")

    w(f"\n{bar}\n  THREAT & ANOMALY FINDINGS ({len(findings)})\n{bar}\n")
    if not findings:
        w("  No suspicious patterns detected with current thresholds.\n")
    for i, f in enumerate(findings, 1):
        w(f"\n  [{i}] ({f.severity}/{f.confidence} confidence) {f.title}\n")
        w(_wrap(f.description, "      "))
        w("\n      -> Recommendation: ")
        w(_wrap(f.recommendation, "         ").lstrip())
        w("\n")

    w(f"\n{bar}\n  SUGGESTED RULES ({rule_fmt} syntax)\n{bar}\n")
    w(f"\n  Subnet aggregation: {result.config.subnet_mode}")
    if result.config.subnet_mode == "adaptive":
        w(f" (>= {result.config.subnet_min_hosts} hosts/subnet collapse to "
          f"/{result.config.prefix_v4})")
    w("\n\n  -- BLOCK rules (from findings) --\n")
    if result.block_rules:
        for r in result.block_rules:
            w(f"  {render_rule(r, rule_fmt)}\n")
    else:
        w("  (none)\n")
    w("\n  -- ALLOW rules (top observed legitimate flows) --\n")
    for r in result.allow_rules:
        w(f"  {render_rule(r, rule_fmt)}\n")
    w("\n")
    return "".join(out)


def print_report(result: "AnalysisResult") -> None:
    sys.stdout.write(format_report(result))


def _wrap(text: str, indent: str, width: int = 70) -> str:
    import textwrap
    return "\n".join(textwrap.wrap(text, width=width,
                                   initial_indent=indent,
                                   subsequent_indent=indent))


# ---------------------------------------------------------------------------
# Optional outputs
# ---------------------------------------------------------------------------

def write_json(path, result: "AnalysisResult") -> None:
    mapping = result.mapping
    summary = result.summary
    findings = result.findings
    allow_rules = result.allow_rules
    block_rules = result.block_rules
    rule_fmt = result.config.rule_format

    def rule_dict(r: Rule) -> Dict[str, Any]:
        return {"decision": r.decision, "src": r.src, "dst": r.dst,
                "port": r.port, "proto": r.proto, "reason": r.reason,
                "rendered": render_rule(r, rule_fmt)}

    payload = {
        "tool": "firewall_analyzer",
        "version": __version__,
        "generated": datetime.now().isoformat(timespec="seconds"),
        "detected_schema": mapping,
        "summary": {k: v for k, v in summary.items() if k != "timespan"},
        "timespan": ([t.isoformat() for t in summary["timespan"]]
                     if summary["timespan"] else None),
        "findings": [
            {"title": f.title, "severity": f.severity,
             "confidence": f.confidence, "description": f.description,
             "recommendation": f.recommendation, "entities": f.entities}
            for f in findings
        ],
        "block_rules": [rule_dict(r) for r in block_rules],
        "allow_rules": [rule_dict(r) for r in allow_rules],
    }
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, default=str)
    log.info("Wrote JSON report: %s", path)


def write_excel(path, result: "AnalysisResult") -> bool:
    summary = result.summary
    findings = result.findings
    allow_rules = result.allow_rules
    block_rules = result.block_rules
    rule_fmt = result.config.rule_format
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log.warning("openpyxl not installed; skipping Excel output. "
                    "Install with: pip install openpyxl")
        return False

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def autosize(ws):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 60)

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("Metric", "Value"),
        ("Records analyzed", summary["total"]),
        ("Unique sources", summary["unique_sources"]),
        ("Unique targets", summary["unique_targets"]),
        ("Allowed", summary["allowed"]),
        ("Blocked", summary["blocked"]),
        ("External inbound", summary["external_inbound"]),
        ("Internal outbound", summary["internal_outbound"]),
    ]
    for r in rows:
        ws.append(r)
    style_header(ws, 2)
    autosize(ws)

    # Top ports
    ws = wb.create_sheet("Top Ports")
    ws.append(("Destination Port", "Service", "Connection Count"))
    for port, count in summary["top_dst_ports"]:
        ws.append((port, service_for_port(port), count))
    style_header(ws, 3)
    autosize(ws)

    # Top talkers
    ws = wb.create_sheet("Top Sources")
    ws.append(("Source IP", "Scope", "Connection Count"))
    for ip, count in summary["top_talkers"]:
        scope = "internal" if _is_internal(_parse_ip(ip)) else "external"
        ws.append((ip, scope, count))
    style_header(ws, 3)
    autosize(ws)

    # Findings
    ws = wb.create_sheet("Findings")
    ws.append(("Severity", "Confidence", "Title", "Description",
               "Recommendation"))
    sev_fill = {"CRITICAL": "C00000", "HIGH": "E26B0A", "MEDIUM": "FFC000",
                "LOW": "FFFF00", "INFO": "D9D9D9"}
    for f in findings:
        ws.append((f.severity, f.confidence, f.title, f.description,
                   f.recommendation))
        ws.cell(row=ws.max_row, column=1).fill = PatternFill(
            "solid", fgColor=sev_fill.get(f.severity, "FFFFFF"))
    style_header(ws, 5)
    for col_letter, width in (("A", 12), ("B", 12), ("C", 38),
                              ("D", 70), ("E", 60)):
        ws.column_dimensions[col_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Rules
    ws = wb.create_sheet("Suggested Rules")
    ws.append(("Decision", "Source", "Destination", "Port", "Protocol",
               f"{rule_fmt} syntax", "Reason"))
    for r in block_rules + allow_rules:
        ws.append((r.decision, r.src, r.dst,
                   r.port if r.port is not None else "any",
                   r.proto, render_rule(r, rule_fmt).split("    #")[0].strip(),
                   r.reason))
    style_header(ws, 7)
    autosize(ws)

    wb.save(path)
    log.info("Wrote Excel workbook: %s", path)
    return True


def write_charts(outdir, result: "AnalysisResult") -> bool:
    summary = result.summary
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except ImportError:
        log.warning("matplotlib not installed; skipping charts. "
                    "Install with: pip install matplotlib")
        return False

    os.makedirs(outdir, exist_ok=True)

    # Top destination ports
    if summary["top_dst_ports"]:
        ports = [str(p) for p, _ in summary["top_dst_ports"][:10]]
        counts = [c for _, c in summary["top_dst_ports"][:10]]
        plt.figure(figsize=(9, 5))
        plt.bar(ports, counts, color="#305496")
        plt.title("Top Destination Ports")
        plt.xlabel("Destination port")
        plt.ylabel("Connections")
        plt.tight_layout()
        plt.savefig(os.path.join(outdir, "top_ports.png"), dpi=120)
        plt.close()

    # Allow vs block
    if summary["has_action"] and (summary["allowed"] or summary["blocked"]):
        plt.figure(figsize=(5, 5))
        plt.pie([summary["allowed"], summary["blocked"]],
                labels=["Allowed", "Blocked"], autopct="%1.1f%%",
                colors=["#70AD47", "#C00000"])
        plt.title("Allowed vs Blocked")
        plt.tight_layout()
        plt.savefig(os.path.join(outdir, "allow_vs_block.png"), dpi=120)
        plt.close()

    # Top talkers
    if summary["top_talkers"]:
        ips = [ip for ip, _ in summary["top_talkers"][:10]]
        counts = [c for _, c in summary["top_talkers"][:10]]
        plt.figure(figsize=(9, 5))
        plt.barh(ips[::-1], counts[::-1], color="#E26B0A")
        plt.title("Top Source IPs")
        plt.xlabel("Connections")
        plt.tight_layout()
        plt.savefig(os.path.join(outdir, "top_talkers.png"), dpi=120)
        plt.close()

    log.info("Wrote charts to: %s", outdir)
    return True


def write_pdf(path, result: "AnalysisResult") -> bool:
    """Render a readable PDF report using reportlab (optional dependency)."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle)
    except ImportError:
        log.warning("reportlab not installed; skipping PDF output. "
                    "Install with: pip install reportlab")
        return False

    summary = result.summary
    styles = getSampleStyleSheet()
    small = ParagraphStyle("small", parent=styles["BodyText"], fontSize=8,
                           leading=10)
    sev_color = {"CRITICAL": colors.HexColor("#C00000"),
                 "HIGH": colors.HexColor("#E26B0A"),
                 "MEDIUM": colors.HexColor("#BF9000"),
                 "LOW": colors.HexColor("#808000"),
                 "INFO": colors.grey}

    doc = SimpleDocTemplate(path, pagesize=letter,
                            leftMargin=0.7 * inch, rightMargin=0.7 * inch,
                            topMargin=0.7 * inch, bottomMargin=0.7 * inch)
    story = []
    story.append(Paragraph("Firewall Log Analysis Report", styles["Title"]))
    story.append(Paragraph(
        f"Generated {datetime.now():%Y-%m-%d %H:%M} &middot; "
        f"firewall_analyzer v{__version__}", styles["Normal"]))
    story.append(Spacer(1, 12))

    # Summary table
    summary_rows = [
        ["Records analyzed", summary["total"]],
        ["Unique sources", summary["unique_sources"]],
        ["Unique targets", summary["unique_targets"]],
    ]
    if summary["has_action"]:
        summary_rows.append(["Allowed / Blocked",
                             f"{summary['allowed']} / {summary['blocked']}"])
    summary_rows.append(["External inbound", summary["external_inbound"]])
    summary_rows.append(["Internal outbound", summary["internal_outbound"]])
    t = Table([["Metric", "Value"]] + summary_rows, colWidths=[200, 200])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    story.append(t)
    story.append(Spacer(1, 14))

    # Findings
    story.append(Paragraph(
        f"Threat &amp; Anomaly Findings ({len(result.findings)})",
        styles["Heading2"]))
    if not result.findings:
        story.append(Paragraph("No suspicious patterns detected with current "
                               "thresholds.", styles["BodyText"]))
    for i, f in enumerate(result.findings, 1):
        head = ParagraphStyle(f"f{i}", parent=styles["Heading4"],
                              textColor=sev_color.get(f.severity, colors.black),
                              spaceBefore=8, spaceAfter=2)
        story.append(Paragraph(f"[{i}] {f.severity} &middot; {f.title}", head))
        story.append(Paragraph(f.description, small))
        story.append(Paragraph(f"<b>Recommendation:</b> {f.recommendation}",
                               small))

    # Rules
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        f"Suggested Rules ({result.config.rule_format} syntax)",
        styles["Heading2"]))

    def rule_table(title, rules):
        story.append(Paragraph(title, styles["Heading4"]))
        if not rules:
            story.append(Paragraph("(none)", small))
            return
        data = [["Rule", "Reason"]]
        for r in rules:
            rendered = render_rule(r, result.config.rule_format).split("    #")[0]
            data.append([Paragraph(rendered, small),
                         Paragraph(r.reason, small)])
        tbl = Table(data, colWidths=[330, 170])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 8))

    rule_table("BLOCK rules (from findings)", result.block_rules)
    rule_table("ALLOW rules (top observed legitimate flows)",
               result.allow_rules)

    doc.build(story)
    log.info("Wrote PDF report: %s", path)
    return True


# ---------------------------------------------------------------------------
# Top-level analysis entry point (shared by CLI and GUI)
# ---------------------------------------------------------------------------

def analyze(csv_path: str, config: Optional[AnalysisConfig] = None
            ) -> AnalysisResult:
    """Run the full pipeline on a CSV and return a structured result.

    This is the single entry point used by both the command line and the GUI.
    """
    config = config or AnalysisConfig()
    records, mapping, _headers = load_records(csv_path)
    if not records:
        raise ValueError("No usable records (need at least source and "
                         "destination IP).")
    summary = summarize(records)
    findings = detect_threats(
        records,
        scan_port_threshold=config.scan_port_threshold,
        sweep_host_threshold=config.sweep_host_threshold,
        block_threshold=config.block_threshold,
        block_ratio=config.block_ratio,
    )
    allow_rules = build_allow_rules(records, findings, config)
    block_rules = build_block_rules(findings, config)
    return AnalysisResult(records=records, mapping=mapping, summary=summary,
                          findings=findings, allow_rules=allow_rules,
                          block_rules=block_rules, config=config)

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="firewall_analyzer.py",
        description="Analyze a firewall/connection-log CSV, detect suspicious "
                    "traffic, and propose firewall rules.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        epilog="Example:\n"
               "  python3 firewall_analyzer.py logs.csv --excel --charts "
               "--rule-format pf\n",
    )
    p.add_argument("csv_file", help="Path to the firewall log CSV.")
    p.add_argument("-o", "--output-dir", default=".",
                   help="Directory for generated files.")
    p.add_argument("--rule-format", choices=["generic", "iptables", "pf"],
                   default="generic", help="Syntax for suggested rules.")
    p.add_argument("--excel", action="store_true",
                   help="Also write a multi-sheet .xlsx report (needs openpyxl).")
    p.add_argument("--pdf", action="store_true",
                   help="Also write a PDF report (needs reportlab).")
    p.add_argument("--json", action="store_true",
                   help="Also write a machine-readable JSON report.")
    p.add_argument("--charts", action="store_true",
                   help="Also write PNG charts (needs matplotlib).")
    # Subnet-based rule aggregation
    p.add_argument("--subnet-mode", choices=["adaptive", "always", "never"],
                   default="adaptive",
                   help="ALLOW-rule aggregation: 'adaptive' collapses a subnet "
                        "only when enough hosts share it; 'always' collapses "
                        "every flow to its subnet; 'never' is one rule per host.")
    p.add_argument("--subnet-min-hosts", type=int, default=3,
                   help="Adaptive mode: distinct hosts in a subnet (to the same "
                        "dst+port) before collapsing to a single CIDR rule.")
    p.add_argument("--subnet-prefix-v4", type=int, default=24,
                   help="IPv4 prefix length used when aggregating to subnets.")
    p.add_argument("--subnet-prefix-v6", type=int, default=64,
                   help="IPv6 prefix length used when aggregating to subnets.")
    p.add_argument("--aggregate-allow", action="store_true",
                   help="Deprecated alias for --subnet-mode always.")
    p.add_argument("--top-rules", type=int, default=25,
                   help="Max number of ALLOW rules to propose.")
    # Detection thresholds
    p.add_argument("--scan-port-threshold", type=int, default=15,
                   help="Distinct dst ports from one src to one dst = port scan.")
    p.add_argument("--sweep-host-threshold", type=int, default=10,
                   help="Distinct dst hosts from one src on one port = sweep.")
    p.add_argument("--block-threshold", type=int, default=5,
                   help="Min blocked events to flag a noisy source.")
    p.add_argument("--block-ratio", type=float, default=0.5,
                   help="Min blocked/total ratio to flag a noisy source.")
    p.add_argument("--gui", action="store_true",
                   help="Launch the graphical interface instead of the CLI.")
    p.add_argument("-q", "--quiet", action="store_true",
                   help="Suppress the console report (still writes files).")
    p.add_argument("-v", "--verbose", action="store_true",
                   help="Verbose logging.")
    p.add_argument("--version", action="version",
                   version=f"%(prog)s {__version__}")
    return p


def config_from_args(args) -> AnalysisConfig:
    mode = "always" if args.aggregate_allow else args.subnet_mode
    return AnalysisConfig(
        rule_format=args.rule_format,
        subnet_mode=mode,
        subnet_min_hosts=args.subnet_min_hosts,
        prefix_v4=args.subnet_prefix_v4,
        prefix_v6=args.subnet_prefix_v6,
        top_rules=args.top_rules,
        scan_port_threshold=args.scan_port_threshold,
        sweep_host_threshold=args.sweep_host_threshold,
        block_threshold=args.block_threshold,
        block_ratio=args.block_ratio,
    )


def main(argv: Optional[Sequence[str]] = None) -> int:
    # Allow `firewall_analyzer.py --gui` to open the graphical front-end.
    if argv is None:
        argv = sys.argv[1:]
    if "--gui" in argv:
        try:
            import firewall_analyzer_gui
        except ImportError as exc:
            print(f"Could not start GUI: {exc}\n"
                  "Ensure firewall_analyzer_gui.py is in the same folder and "
                  "that Tk is available (Linux: install python3-tk).",
                  file=sys.stderr)
            return 2
        return firewall_analyzer_gui.launch()

    args = build_arg_parser().parse_args(argv)
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(message)s",
    )

    if not os.path.isfile(args.csv_file):
        log.error("File not found: %s", args.csv_file)
        return 2

    config = config_from_args(args)
    try:
        result = analyze(args.csv_file, config)
    except (ValueError, OSError) as exc:
        log.error("Failed to analyze CSV: %s", exc)
        return 2

    log.info("Loaded %d records from %s", result.summary["total"],
             args.csv_file)

    if not args.quiet:
        print_report(result)

    os.makedirs(args.output_dir, exist_ok=True)
    base = os.path.splitext(os.path.basename(args.csv_file))[0]

    if args.json:
        write_json(os.path.join(args.output_dir, f"{base}_report.json"), result)
    if args.excel:
        write_excel(os.path.join(args.output_dir, f"{base}_report.xlsx"), result)
    if args.pdf:
        write_pdf(os.path.join(args.output_dir, f"{base}_report.pdf"), result)
    if args.charts:
        write_charts(os.path.join(args.output_dir, f"{base}_charts"), result)

    # Exit code reflects worst finding severity (useful for automation/CI).
    if any(f.severity in ("CRITICAL", "HIGH") for f in result.findings):
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
